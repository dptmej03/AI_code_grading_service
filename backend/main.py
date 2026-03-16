import os
import json
import uuid
import asyncio
import io
from pathlib import Path
from datetime import timedelta
from typing import Dict, Optional

from fastapi import (
    FastAPI, Depends, HTTPException, status, UploadFile, File,
    Form, BackgroundTasks
)
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv

load_dotenv()

from auth import authenticate_user, create_access_token, get_current_user, ACCESS_TOKEN_EXPIRE_MINUTES
from schemas import (
    Token, LoginRequest, GradingCriteria, GradingSession, StudentResult
)
from services.notebook_service import (
    extract_notebooks_from_zip, parse_student_id_from_filename
)
from services.grading_service import grade_student_notebook

app = FastAPI(title="Jupyter Notebook 자동 채점 시스템", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# In-memory session store
grading_sessions: Dict[str, GradingSession] = {}


# ─── Auth ─────────────────────────────────────────────────────────────────────

@app.post("/auth/login", response_model=Token)
async def login(request: LoginRequest):
    user = authenticate_user(request.username, request.password)
    if not user:
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="아이디 또는 비밀번호가 올바르지 않습니다",
            headers={"WWW-Authenticate": "Bearer"},
        )
    token = create_access_token(
        data={"sub": user["username"]},
        expires_delta=timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    )
    return Token(access_token=token, token_type="bearer")


@app.get("/auth/me")
async def get_me(current_user=Depends(get_current_user)):
    return {"username": current_user["username"], "role": current_user["role"]}


# ─── Grading ──────────────────────────────────────────────────────────────────

@app.post("/grading/start")
async def start_grading(
    background_tasks: BackgroundTasks,
    answer_notebook: UploadFile = File(...),
    student_zip: UploadFile = File(...),
    criteria_file: UploadFile = File(...),
    current_user=Depends(get_current_user)
):
    """채점 시작: 파일 업로드 후 백그라운드에서 채점 실행."""
    # Read files
    answer_bytes = await answer_notebook.read()
    zip_bytes = await student_zip.read()
    criteria_bytes = await criteria_file.read()

    # Parse criteria
    try:
        criteria_data = json.loads(criteria_bytes.decode('utf-8'))
        criteria = GradingCriteria(**criteria_data)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"채점 기준 파일 파싱 오류: {str(e)}")

    # Extract student notebooks
    try:
        student_notebooks = extract_notebooks_from_zip(zip_bytes)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"ZIP 파일 처리 오류: {str(e)}")

    if not student_notebooks:
        raise HTTPException(status_code=400, detail="ZIP 파일 내에 .ipynb 파일이 없습니다")

    session_id = str(uuid.uuid4())
    session = GradingSession(
        session_id=session_id,
        status="pending",
        progress=0.0,
        total_students=len(student_notebooks),
        processed_students=0,
        results=[]
    )
    grading_sessions[session_id] = session

    background_tasks.add_task(
        run_grading_session,
        session_id, answer_bytes, student_notebooks, criteria
    )

    return {"session_id": session_id, "total_students": len(student_notebooks)}


async def run_grading_session(
    session_id: str,
    answer_bytes: bytes,
    student_notebooks: list,
    criteria: GradingCriteria
):
    session = grading_sessions[session_id]
    session.status = "running"
    total = len(student_notebooks)

    for i, (filename, content) in enumerate(student_notebooks):
        session.current_student = filename
        try:
            problem_results, error = await grade_student_notebook(
                student_nb_content=content,
                answer_nb_content=answer_bytes,
                criteria=criteria,
                execute=False
            )
            total_score = sum(p.obtained_score for p in problem_results)
            max_total = sum(p.full_score for p in problem_results)

            student_result = StudentResult(
                filename=filename,
                student_id=parse_student_id_from_filename(filename),
                total_score=total_score,
                max_total_score=max_total,
                problems=problem_results,
                error=error
            )
            session.results.append(student_result)
        except Exception as e:
            session.results.append(StudentResult(
                filename=filename,
                student_id=parse_student_id_from_filename(filename),
                total_score=0,
                max_total_score=sum(p.full_score for p in criteria.problems),
                problems=[],
                error=str(e)
            ))

        session.processed_students = i + 1
        session.progress = ((i + 1) / total) * 100

    session.status = "completed"
    session.progress = 100.0
    session.current_student = None


@app.get("/grading/session/{session_id}")
async def get_session(session_id: str, current_user=Depends(get_current_user)):
    session = grading_sessions.get(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="세션을 찾을 수 없습니다")
    return session


@app.get("/grading/session/{session_id}/results")
async def get_results(session_id: str, current_user=Depends(get_current_user)):
    session = grading_sessions.get(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="세션을 찾을 수 없습니다")
    if session.status != "completed":
        raise HTTPException(status_code=400, detail="채점이 아직 완료되지 않았습니다")
    return session.results


@app.get("/grading/session/{session_id}/download")
async def download_excel(session_id: str, current_user=Depends(get_current_user)):
    """채점 결과를 Excel 파일로 다운로드."""
    session = grading_sessions.get(session_id)
    if not session or session.status != "completed":
        raise HTTPException(status_code=400, detail="채점이 완료되지 않았습니다")

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "채점결과"

    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Collect all problem IDs
    all_problem_ids = []
    if session.results:
        for r in session.results:
            for p in r.problems:
                if p.problem_id not in all_problem_ids:
                    all_problem_ids.append(p.problem_id)
    all_problem_ids.sort()

    # Build headers
    headers = ["학번/이름", "파일명"]
    for pid in all_problem_ids:
        problem = next((p for r in session.results for p in r.problems if p.problem_id == pid), None)
        max_s = problem.full_score if problem else 0
        headers.append(f"문제{pid} ({max_s}점)")
    headers.extend(["총점", "만점", "비율(%)"])

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Data rows
    for row_idx, student in enumerate(session.results, 2):
        ws.cell(row=row_idx, column=1, value=student.student_id)
        ws.cell(row=row_idx, column=2, value=student.filename)
        col = 3
        for pid in all_problem_ids:
            p = next((p for p in student.problems if p.problem_id == pid), None)
            ws.cell(row=row_idx, column=col, value=p.obtained_score if p else 0)
            col += 1
        ws.cell(row=row_idx, column=col, value=student.total_score)
        ws.cell(row=row_idx, column=col+1, value=student.max_total_score)
        ratio = (student.total_score / student.max_total_score * 100) if student.max_total_score > 0 else 0
        ws.cell(row=row_idx, column=col+2, value=round(ratio, 1))
        for c in range(1, len(headers)+1):
            ws.cell(row=row_idx, column=c).border = thin_border
            ws.cell(row=row_idx, column=c).alignment = center_align

    # Auto column width
    for col in range(1, len(headers)+1):
        max_len = max(
            len(str(ws.cell(row=r, column=col).value or ""))
            for r in range(1, len(session.results)+2)
        )
        ws.column_dimensions[get_column_letter(col)].width = max(max_len + 4, 12)

    # Detail sheet
    ws2 = wb.create_sheet("상세채점")
    detail_headers = ["학번/이름", "문제", "채점항목", "최대점수", "획득점수", "피드백"]
    for col, h in enumerate(detail_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    row = 2
    for student in session.results:
        for problem in student.problems:
            for ps in problem.partial_scores:
                ws2.cell(row=row, column=1, value=student.student_id)
                ws2.cell(row=row, column=2, value=f"문제{problem.problem_id}")
                ws2.cell(row=row, column=3, value=ps.item)
                ws2.cell(row=row, column=4, value=ps.max_score)
                ws2.cell(row=row, column=5, value=ps.score)
                ws2.cell(row=row, column=6, value=ps.reason)
                for c in range(1, 7):
                    ws2.cell(row=row, column=c).border = thin_border
                row += 1

    for col in range(1, 7):
        max_len = max(
            len(str(ws2.cell(r2, col).value or "")) for r2 in range(1, row)
        )
        ws2.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=grading_results_{session_id[:8]}.xlsx"}
    )


@app.get("/health")
async def health():
    return {"status": "ok"}
